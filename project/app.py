import os
import requests
import urllib.parse
import openpyxl
from cs50 import SQL
from flask import Flask, flash, redirect, render_template, request, session, make_response
from flask_session import Session
from tempfile import mkdtemp
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
from datetime import datetime, timedelta
import time


app = Flask(__name__)

# Ensure templates are auto-reloaded
app.config["TEMPLATES_AUTO_RELOAD"] = True

# Configure session to use file system (insted of signed cookies)
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

# Configure SQLite database
db = SQL("sqlite:///database.db")

vat = 1.2

@app.after_request
def after_request(response):
    """ Ensure responses are not cached"""
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Expires"] = 0
    response.headers["Pragma"] = "no-cache"
    return response
    

def login_required(f):
    """ Decorate routes to require login. """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("user_id") is None:
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated_function

# Log In , Log Out, Register --------------------
@app.route("/register", methods=["GET", "POST"])
def register():
    """Register user"""
    if request.method == "GET":
        return render_template("register.html")

    # Ensure username was submitted
    if not request.form.get("username"):
        errmsg = 1
        flash("Username not provided")
        return render_template("register.html", errmsg=errmsg)
    # Ensure password was submitted
    elif not request.form.get("password"):
        errmsg = 1
        flash("Password not provided")
        return render_template("register.html", errmsg=errmsg)
    # Ensure password confirmation was submitted
    elif not request.form.get("password2"):
        errmsg = 1
        flash("Password confirmation not provided")
        return render_template("register.html", errmsg=errmsg)

    username = request.form.get("username")
    password = request.form.get("password")
    confirmation = request.form.get("password2")
    date = datetime.now()
    
    # Ensure username is not in database
    usernames = db.execute("SELECT username FROM users")
    names = []
    for i in range(len(usernames)):
        names.append(usernames[i]['username'])

    if username in names:
        errmsg = 1
        flash("Username exists, please choose a different username")
        return render_template("register.html", errmsg=errmsg)

    # Ensure passwords are match
    if password != confirmation:
        errmsg = 1
        flash("Passwords do not match")
        return render_template("register.html", errmsg=errmsg)


    hashed_pass = generate_password_hash(password, method='pbkdf2:sha256', salt_length=8)

    db.execute("INSERT INTO users(username, password, date) VALUES(?,?,?)", username, hashed_pass, date)
    errmsg = 0
    flash("Registration is successful, please log in")
    return render_template("register.html", errmsg=errmsg)  

@app.route("/login", methods=["GET", "POST"])
def login():
    # forget any user id
    session.clear()

    # User reached route via POST (as by submittin a form via POST)
    if request.method == "POST":
        # ensure username submitted
        if not request.form.get("username") or not request.form.get("password"):
            errmsg = 1
            flash("Please fill all empty fields")
            return render_template("login.html", errmsg=errmsg)    
        
        # check database for username
        rows = db.execute("SELECT * FROM users WHERE username = ?", request.form.get("username"))

        # Ensure username exists and password is correct
        if len(rows) != 1 or not check_password_hash(rows[0]["password"], request.form.get("password")):
            errmsg = 1
            flash("Username or Password is not correct!")
            return render_template("login.html", errmsg=errmsg)

         # Remember which user has logged in
        session["user_id"] = rows[0]["user_id"]
        session["username"] = rows[0]["username"]

        # Redirect user to home page
        return redirect("/")   

      # User reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("login.html")

@app.route("/logout")
def logout():
    """Log user out"""

    # Forget any user_id
    session.clear()

    # Redirect user to login form
    return redirect("/")

# Main Menu --------------------

@app.route("/")
@login_required
def index():
    user_id = session.get("user_id")
    list = db.execute("SELECT * from data WHERE user_id=?", user_id)
    if len(list) != 0:
        return portfolio()
    else:
        return newproject()

@app.route("/newproject", methods=["GET", "POST"])
@login_required
def newproject():

    if request.method == "POST":
        # Ensure fields are not empty
        if not request.form.get("p_name") or not request.form.get("p_start"):
            errmsg = 1
            flash("Please fill all empty fields")
            return render_template("newproject.html", errmsg=errmsg)


        user_id = session.get("user_id")
        start = request.form.get("p_start")
        project_name = request.form.get("p_name")
        date = datetime.now()

        # Ensure if project does not already exists
        list = db.execute("SELECT project_name FROM projects WHERE project_name = ?", project_name)
        if len(list) > 0:
            errmsg = 1
            flash("Project name already exists in database")
            return render_template("newproject.html", errmsg=errmsg)

        # Insert new row for new project
        db.execute("INSERT INTO projects(user_id, project_name, project_start, date) VALUES(?,?,?,?)",\
             user_id, project_name, start, date)
        errmsg = 0
        flash("New project created. Initiate update to start updating")
        projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)
        return render_template("uinit.html", projects=projects, errmsg=errmsg)

    return render_template("newproject.html")

@app.route("/portfolio", methods=["GET", "POST"])
@login_required
def portfolio():

    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)
    dates = db.execute("SELECT DISTINCT cutoff FROM data WHERE user_id=?", user_id)
    list = db.execute("SELECT * from data WHERE user_id=?", user_id)
    
    # If new user has no project yet
    if len(projects) == 0:
        errmsg = 1
        flash("Please create a new project and update it first")
        return render_template("newproject.html", errmsg=errmsg)

    
    # If new user has no data yet
    try:
        if len(list) == 0:
            errmsg = 2
            flash("Please update new project first")
            return render_template("warning.html", errmsg=errmsg, projects=projects)
    except TypeError:
        errmsg = 2
        flash("No data to show yet. Please update new project first")
        return render_template("warning.html", errmsg=errmsg, projects=projects)

    # Default date to show on portfoli page
    max_date = db.execute("SELECT MAX (cutoff) from data WHERE user_id=?", user_id)[0]["MAX (cutoff)"]

    if request.method == "POST":
        
        cutoff = request.form.get("cutoff")
        data = db.execute("SELECT * FROM data WHERE user_id=? AND cutoff=?", user_id, cutoff)
        total_income = db.execute("SELECT SUM (income) FROM data WHERE user_id=? AND cutoff=?", user_id, cutoff)[0]["SUM (income)"]
        total_expense = db.execute("SELECT SUM (expense) FROM data WHERE user_id=? AND cutoff=?", user_id, cutoff)[0]["SUM (expense)"]
        total_profit = total_income - total_expense

        return render_template("portfolio.html", projects=projects, dates=dates, data=data, total_income=total_income,\
        total_profit=total_profit, cutoff=cutoff)

    cutoff = max_date
    data = db.execute("SELECT * FROM data WHERE user_id=? AND cutoff=?", user_id, cutoff)
    
    total_income = db.execute("SELECT SUM (income) FROM data WHERE user_id=? AND cutoff=?", user_id, cutoff)[0]["SUM (income)"]
    total_expense = db.execute("SELECT SUM (expense) FROM data WHERE user_id=? AND cutoff=?", user_id, cutoff)[0]["SUM (expense)"]
    total_profit = total_income - total_expense
    
    
    return render_template("portfolio.html", projects=projects, dates=dates, data=data, total_income=total_income,\
    total_profit=total_profit, max_date=max_date, cutoff=cutoff)


@app.route("/dashboard", methods=["GET", "POST"])
@login_required
def dashboard():
    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)
    dates = db.execute("SELECT DISTINCT cutoff FROM data WHERE user_id=?", user_id)
    list = db.execute("SELECT * from data WHERE user_id=?", user_id)
    
    # If new user has no project yet
    try:
        if len(list) == 0:
            errmsg = 2
            flash("Please update new project first")
            return render_template("warning.html", errmsg=errmsg, projects=projects)
    except TypeError:
        errmsg = 2
        flash("No data to show yet. Please update new project first")
        return render_template("warning.html", errmsg=errmsg, projects=projects)

    # Default date to show on dashboard
    #max_date = dates[len(dates) - 1]["cutoff"]
    max_date = db.execute("SELECT MAX (cutoff) from data WHERE user_id=?", user_id)[0]["MAX (cutoff)"]

    if request.method == "POST":
        
        cutoff = request.form.get("cutoff")
        project_name = request.form.get("projectname")
        data = db.execute("SELECT * FROM data WHERE user_id=? AND cutoff=? AND project_name=?", user_id, cutoff, project_name)
        
        try:
            # Values for Progress Pie Charts
            actual_progress = '%.1f' % (round(data[0]['earned_mh'] / data[0]['total_mh'] , 3) * 100)
            if data[0]['earned_mh'] / data[0]['total_mh'] > 1:
                actual_progress = 100

            planned_progress = '%.1f' % (data[0]['planned_mh'] / data[0]['total_mh'] * 100)
            if data[0]['planned_mh'] / data[0]['total_mh'] > 1:
                planned_progress = 100

            start = db.execute("SELECT project_start FROM projects WHERE user_id=? AND project_name=?", user_id, project_name)[0]['project_start']
            finish = data[0]['contract_finish']
            finishp = data[0]['planned_finish']

            
            # Duration passed
            durpass = datetime.strptime(cutoff, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')
            # Total Duration
            duration = datetime.strptime(finish, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')
            # Duration passed percentage
            durpass_perc = '%.1f' % (round(durpass / duration, 3 ) * 100)
            if durpass / duration > 1:
                durpass = 100

            duration_m = (duration / (365/12)).days

            # Values for Manhour and CPI-SPI Section
            total_mh = f"{(data[0]['total_mh']):,.0f}".replace(",", " ")
            earned_mh = f"{(data[0]['earned_mh']):,.0f}".replace(",", " ")
            estimated_mh = f"{(data[0]['estimated_mh']):,.0f}".replace(",", " ")
            actual_mh = f"{(data[0]['actual_mh']):,.0f}".replace(",", " ")
            cpi = f"{(data[0]['actual_mh'] / data[0]['earned_mh']):.2f}"
            spi = f"{(data[0]['earned_mh'] / data[0]['planned_mh']):.2f}"

            # Values Budget Numbers Section
            income = f"{(data[0]['income']):,.0f}".replace(",", " ")
            expense = f"{(data[0]['expense']):,.0f}".replace(",", " ")
            profit = f"{(data[0]['income'] - data[0]['expense']):,.0f}".replace(",", " ")
            profitrate = '%.1f' % (round( (data[0]['income'] - data[0]['expense']) / data[0]['income'], 3 ) * 100)

            # Values for Budgte Graph
            income_m = round((data[0]['income'] / 1000000), 1)
            expense_m = round((data[0]['expense'] / 1000000), 1)
            profit_m = round((data[0]['income'] - data[0]['expense']) / 1000000, 1)

            # Values for Financials
            invoice = f"{(data[0]['invoice']):,.0f}".replace(",", " ")
            cash = f"{(data[0]['cash']):,.0f}".replace(",", " ")
            cash_balance = f"{(data[0]['cash_balance']):,.0f}".replace(",", " ")
            accrual_balance = f"{(data[0]['accrual_balance']):,.0f}".replace(",", " ")
            invoice_p = round((data[0]['invoice'] / data[0]['income']) * 100, 1)
            cash_p = round((data[0]['cash'] / (data[0]['income'] * vat)) * 100, 1)

            # Values for Financial Graph
            invoice_m = round((data[0]['invoice'] / 1000000), 1)
            cash_m = round((data[0]['cash'] / 1000000), 1)
        
        except IndexError:
            errmsg = 2
            flash("No data to show. Please change your selection")
            data = db.execute("SELECT * FROM data WHERE user_id=? AND cutoff=?", user_id, cutoff)
            return render_template("dashboard.html",  projects=projects, dates=dates, max_date=max_date, cutoff=cutoff, project_name=project_name, errmsg=errmsg)


        
        return render_template("dashboard.html", projects=projects, dates=dates, cutoff=cutoff, project_name=project_name,\
            actual_progress=actual_progress, planned_progress=planned_progress, durpass_perc=durpass_perc,\
            total_mh=total_mh, earned_mh=earned_mh, estimated_mh=estimated_mh, actual_mh=actual_mh, cpi=cpi, spi=spi,\
            income=income, expense=expense, profit=profit, profitrate=profitrate, income_m=income_m, expense_m=expense_m, profit_m=profit_m,\
            invoice=invoice, cash=cash, cash_balance=cash_balance, accrual_balance=accrual_balance, invoice_p=invoice_p, cash_p=cash_p,\
            invoice_m=invoice_m, cash_m=cash_m, start=start, finish=finish, finishp=finishp, duration_m=duration_m )


    cutoff = 'Select Date'
    project_name = 'Select Project'
    data = db.execute("SELECT * FROM data WHERE user_id=? AND cutoff=?", user_id, cutoff)

    return render_template("dashboard.html", projects=projects, dates=dates, max_date=max_date, cutoff=cutoff, project_name=project_name)

@app.route("/reports", methods=["GET", "POST"])
@login_required
def reports():
    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)
    dates = db.execute("SELECT DISTINCT cutoff FROM data WHERE user_id=?", user_id)

    if request.method == "POST":

        projects_selected = request.form.getlist("project_names")
        project_count = len(projects_selected)
        cutoff = request.form.get("cutoff")
        report_type = request.form.get("report_type")

        try:
            # Ensure that at all selected
            if project_count == 0 or not request.form.get("cutoff") or not request.form.get("report_type"):
                errmsg = 2
                flash("Please choose 'Report Type', 'Projects' and 'Cut-off Date' to a generate report")
                return render_template("reports.html", projects=projects, dates=dates, errmsg=errmsg )

            table_data = []
            
            for i in range(project_count):
                project_data = {}
                project_data['project_name'] = projects_selected[i]
                data2 = db.execute("SELECT * FROM data WHERE user_id=? AND cutoff=? AND project_name=?", user_id, cutoff, projects_selected[i] )

                # Flash message if there is no data on selected options
                if len(data2) == 0:
                    errmsg = 2
                    flash("No data to show on selected cut-off date")
                    return render_template("reports.html", projects=projects, dates=dates, errmsg=errmsg )
                
                # STATUS REPORT VALUES
                project_data['planned_progress'] = '%.1f' % (data2[0]['planned_mh'] / data2[0]['total_mh'] * 100)
                project_data['actual_progress'] = '%.1f' % (round(data2[0]['earned_mh'] / data2[0]['total_mh'] , 3) * 100)
                project_data['variance'] = '%.1f' % (round((data2[0]['earned_mh'] - data2[0]['planned_mh']) / data2[0]['total_mh'], 3) * 100)

                project_data['planned_mh'] = f"{data2[0]['planned_mh']:,.0f}".replace(",", " ")
                project_data['earned_mh'] = f"{data2[0]['earned_mh']:,.0f}".replace(",", " ")
                project_data['actual_mh'] = f"{data2[0]['actual_mh']:,.0f}".replace(",", " ")
                project_data['cpi'] = f"{data2[0]['actual_mh']:,.0f}".replace(",", " ")
                project_data['actual_mh'] = f"{data2[0]['actual_mh']:,.0f}".replace(",", " ")

                project_data['whitecollar_mp'] = f"{data2[0]['whitecollar_mp']:,.0f}".replace(",", " ")
                project_data['bluecollar_mp'] = f"{data2[0]['bluecollar_mp']:,.0f}".replace(",", " ")
                project_data['subcon_mp'] = f"{data2[0]['subcon_mp']:,.0f}".replace(",", " ")
                project_data['total_mp'] = f"{data2[0]['subcon_mp'] + data2[0]['bluecollar_mp'] + data2[0]['whitecollar_mp']:,.0f}".replace(",", " ")
                project_data['direct_mp_ratio'] = '%.1f' % (round(data2[0]['whitecollar_mp'] / (data2[0]['subcon_mp'] + data2[0]['bluecollar_mp'] + data2[0]['whitecollar_mp']), 3 ) * 100)

                project_data['start'] = db.execute("SELECT project_start FROM projects WHERE user_id=? AND project_name=?", user_id, projects_selected[i])[0]['project_start']
                project_data['contract_finish'] = data2[0]['contract_finish']
                project_data['planned_finish'] = data2[0]['planned_finish']

                # FINANCE REPORT VALUES
                project_data['income'] =f"{data2[0]['income']:,.0f}".replace(",", " ")
                project_data['expense'] =f"{data2[0]['expense']:,.0f}".replace(",", " ")
                project_data['profit'] =f"{data2[0]['income'] - data2[0]['expense']:,.0f}".replace(",", " ")
                project_data['profit_ratio'] = '%.1f' % (((data2[0]['income'] - data2[0]['expense']) / data2[0]['income']) * 100)
                project_data['invoice'] =f"{data2[0]['invoice']:,.0f}".replace(",", " ")
                project_data['cash'] =f"{data2[0]['cash']:,.0f}".replace(",", " ")
                project_data['cash_balance'] =f"{data2[0]['cash_balance']:,.0f}".replace(",", " ")
                project_data['accrual_balance'] =f"{data2[0]['accrual_balance']:,.0f}".replace(",", " ")

                # KPI REPORT VALUES
                project_data['actual_p'] = '%.1f' % ((data2[0]['earned_mh'] / data2[0]['total_mh']) * 100)
                project_data['planned_p'] = '%.1f' % ((data2[0]['planned_mh'] / data2[0]['total_mh']) * 100)
                project_data['invoice_p'] = '%.1f' % ((data2[0]['invoice'] / data2[0]['income']) * 100)
                project_data['cash_p'] = '%.1f' % ((data2[0]['cash'] / (data2[0]['income'] * vat) ) * 100)
                project_data['cpi'] = f"{(data2[0]['actual_mh'] / data2[0]['earned_mh']):.2f}"
                project_data['spi'] = f"{(data2[0]['earned_mh'] / data2[0]['planned_mh']):.2f}"
                
                # Duration passed
                start = db.execute("SELECT project_start FROM projects WHERE user_id=? AND project_name=?", user_id, projects_selected[i])[0]['project_start']
                finish = data2[0]['contract_finish']

                durpass = datetime.strptime(cutoff, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')
                # Total Duration
                duration = datetime.strptime(finish, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')
                # Duration passed percentage
                project_data['duration_p'] = '%.1f' % (round(durpass / duration, 3 ) * 100)
                
                #Add project_data dictionary to table-data list after every loop 
                table_data.append(project_data)

                table_count =  len(table_data)

        except (IndexError):
            errmsg = 2
            flash("No data to show. Please change your selection")
            data = db.execute("SELECT * FROM data WHERE user_id=? AND cutoff=?", user_id, cutoff)
            return render_template("reports.html", projects=projects, dates=dates, cutoff=cutoff,\
                projects_selected=projects_selected, table_data=table_data, project_count=project_count, report_type=report_type, errmsg=errmsg )

        return render_template("reports.html", projects=projects, dates=dates, cutoff=cutoff,\
            projects_selected=projects_selected, table_data=table_data, project_count=project_count, report_type=report_type, table_count=table_count )

    return render_template("reports.html", projects=projects, dates=dates, )

# Update Menu -----------------------------------

@app.route("/uinit", methods=["POST"])
@login_required
def uinit():
    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)

    if request.method == "POST": 
        #Ensure fields are not empty
        if not request.form.get("project_name") or not request.form.get("cutoff"):
            errmsg = 1
            flash("Please fill all empty fields")
            return render_template("uinit.html", errmsg=errmsg, projects=projects)

        project_name = request.form.get("project_name")
        project_id = db.execute("SELECT project_id FROM projects WHERE project_name=?", project_name)[0]['project_id']
        cutoff_date = request.form.get("cutoff")
        input_date = datetime.now()

        
        db.execute("INSERT INTO data (user_id, project_id, project_name, cutoff, date) \
            VALUES(?,?,?,?,?)", user_id, project_id, project_name, cutoff_date, input_date)
        errmsg = 0
        flash("Database initiated for update, please update the project")
        projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)
        return render_template("uschedule.html", errmsg=errmsg, projects=projects)

@app.route("/uschedule", methods=["GET", "POST"])
@login_required
def uschedule():
    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)

    if request.method == "POST":
        # Ensure fields are not empty
        if not request.form.get("project_name") or not request.form.get("cutoff")\
        or not request.form.get("contract_finish") or not request.form.get("planned_finish"):
            errmsg = 1
            flash("Please fill all empty fields")
            return render_template("uschedule.html", errmsg=errmsg, projects=projects)

        project_name = request.form.get("project_name")
        project_id = db.execute("SELECT project_id FROM projects WHERE project_name=?", project_name)[0]['project_id']
        cutoff_date = request.form.get("cutoff")
        contract_finish = request.form.get("contract_finish")
        planned_finish = request.form.get("planned_finish")
        input_date = datetime.now()

        
        list = db.execute("SELECT project_id, cutoff, contract_finish FROM data WHERE project_id=? AND cutoff=?", project_id, cutoff_date)

        # initiate update if there is no data on cut-off date
        if len(list) == 0:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        if list[0]['cutoff'] != cutoff_date:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        # Check if already updated for selected cut-off date
        if list[0]['project_id'] == project_id and list[0]['cutoff'] == cutoff_date and list[0]['contract_finish'] != None:
            errmsg = 1
            flash ("Data already exists on the cut-off date selected")
            return render_template("uschedule.html",errmsg=errmsg, projects=projects)

        # Update entry
        db.execute("UPDATE data SET user_id = ?, contract_finish = ?, planned_finish = ?, date = ? \
            WHERE project_id = ? AND project_name = ? AND cutoff = ?",\
            user_id, contract_finish, planned_finish, input_date, project_id, project_name, cutoff_date)

        errmsg = 0
        flash("Update is successful")
        return render_template("uprogress.html",errmsg=errmsg, projects=projects)
        

    
    return render_template("uschedule.html", projects=projects)

@app.route("/uprogress", methods=["GET", "POST"])
@login_required
def uprogress():
    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)

    if request.method == "POST":
        # Ensure fields are not empty
        if not request.form.get("project_name") or not request.form.get("cutoff")\
        or not request.form.get("total_mh") or not request.form.get("planned_mh")\
        or not request.form.get("earned_mh") or not request.form.get("actual_mh")\
        or not request.form.get("estimated_mh"):
            errmsg = 1
            flash("Please fill all empty fields")
            return render_template("uprogress.html", errmsg=errmsg, projects=projects)

        project_name = request.form.get("project_name")
        project_id = db.execute("SELECT project_id FROM projects WHERE project_name=?", project_name)[0]['project_id']
        cutoff_date = request.form.get("cutoff")
        total_mh = request.form.get("total_mh")
        planned_mh = request.form.get("planned_mh")
        earned_mh = request.form.get("earned_mh")
        actual_mh = request.form.get("earned_mh")
        estimated_mh = request.form.get("estimated_mh")
        input_date = datetime.now()

        list = db.execute("SELECT project_id, cutoff, total_mh FROM data WHERE project_id=? AND cutoff=?", project_id, cutoff_date)

        # initiate update if there is no data on cut-off date
        if len(list) == 0:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        if list[0]['cutoff'] != cutoff_date:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        # Check if already updated for selected cut-off date
        if list[0]['project_id'] == project_id and list[0]['cutoff'] == cutoff_date and list[0]['total_mh'] != None:
            errmsg = 1
            flash ("Data already exists on the cut-off date selected")
            return render_template("uprogress.html",errmsg=errmsg, projects=projects)

        # Update entry
        db.execute("UPDATE data SET user_id = ?, total_mh = ?, planned_mh = ?, earned_mh = ?,\
            actual_mh = ?, estimated_mh = ?, date = ? WHERE project_id = ? AND project_name = ? AND cutoff = ?",\
            user_id, total_mh, planned_mh, earned_mh, actual_mh, estimated_mh, input_date, project_id, project_name, cutoff_date)

        errmsg = 0
        flash("Update is successful")
        return render_template("umanpower.html",errmsg=errmsg, projects=projects)
        

    return render_template("uprogress.html", projects=projects)    

@app.route("/umanpower", methods=["GET", "POST"])
@login_required
def umanpower():
    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)

    if request.method == "POST":
        # Ensure fields are not empty
        if not request.form.get("project_name") or not request.form.get("cutoff")\
        or not request.form.get("bluecollar") or not request.form.get("whitecollar")\
        or not request.form.get("subcon"):
            errmsg = 1
            flash("Please fill all empty fields")
            return render_template("umanpower.html", errmsg=errmsg, projects=projects)


        project_name = request.form.get("project_name")
        project_id = db.execute("SELECT project_id FROM projects WHERE project_name=?", project_name)[0]['project_id']
        cutoff_date = request.form.get("cutoff")
        bluecollar = request.form.get("bluecollar")
        whitecollar = request.form.get("whitecollar")
        subcon = request.form.get("subcon")
        input_date = datetime.now()

        list = db.execute("SELECT project_id, cutoff, bluecollar_mp FROM data WHERE project_id=? AND cutoff=?", project_id, cutoff_date)

        # initiate update if there is no data on cut-off date
        if len(list) == 0:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        if list[0]['cutoff'] != cutoff_date:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        # Check if already updated for selected cut-off date
        if list[0]['project_id'] == project_id and list[0]['cutoff'] == cutoff_date and list[0]['bluecollar_mp'] != None:
            errmsg = 1
            flash ("Data already exists on the cut-off date selected")
            return render_template("umanpower.html",errmsg=errmsg, projects=projects)

        # Update entry
        db.execute("UPDATE data SET user_id = ?, bluecollar_mp = ?, whitecollar_mp = ?, subcon_mp = ?, date = ?\
            WHERE project_id = ? AND project_name = ? AND cutoff = ?",\
            user_id, bluecollar, whitecollar, subcon, input_date, project_id, project_name, cutoff_date)

        errmsg = 0
        flash("Update is successful")
        return render_template("ubudget.html",errmsg=errmsg, projects=projects)
 
    return render_template("umanpower.html", projects=projects) 

@app.route("/ubudget", methods=["GET", "POST"])
@login_required
def ubudget():
    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)

    if request.method == "POST":
        # Ensure fields are not empty
        if not request.form.get("project_name") or not request.form.get("cutoff")\
        or not request.form.get("income") or not request.form.get("expense"):
            errmsg = 1
            flash("Please fill all empty fields")
            return render_template("ubudget.html", errmsg=errmsg, projects=projects)
    
        project_name = request.form.get("project_name")
        project_id = db.execute("SELECT project_id FROM projects WHERE project_name=?", project_name)[0]['project_id']
        cutoff_date = request.form.get("cutoff")
        income = request.form.get("income")
        expense = request.form.get("expense")
        input_date = datetime.now()

        
        list = db.execute("SELECT project_id, cutoff, income FROM data WHERE project_id=? AND cutoff=?", project_id, cutoff_date)

        # initiate update if there is no data on cut-off date
        if len(list) == 0:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        if list[0]['cutoff'] != cutoff_date:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        # Check if already updated for selected cut-off date
        if list[0]['project_id'] == project_id and list[0]['cutoff'] == cutoff_date and list[0]['income'] != None:
            errmsg = 1
            flash ("Data already exists on the cut-off date selected")
            return render_template("ubudget.html",errmsg=errmsg, projects=projects)

        # Update entry
        db.execute("UPDATE data SET user_id = ?, income = ?, expense = ?, date = ? \
            WHERE project_id = ? AND project_name = ? AND cutoff = ?",\
            user_id, income, expense, input_date, project_id, project_name, cutoff_date)

        errmsg = 0
        flash("Update is successful")
        return render_template("ufinance.html",errmsg=errmsg, projects=projects)

    
    return render_template("ubudget.html", projects=projects)

@app.route("/ufinance", methods=["GET", "POST"])
@login_required
def ufinance():
    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)

    if request.method == "POST":
        # Ensure fields are not empty
        if not request.form.get("project_name") or not request.form.get("cutoff")\
        or not request.form.get("invoice") or not request.form.get("accrual_balance")\
        or not request.form.get("cash") or not request.form.get("cash_balance"):
            errmsg = 1
            flash("Please fill all empty fields")
            return render_template("ufinance.html", errmsg=errmsg, projects=projects)
        
        project_name = request.form.get("project_name")
        project_id = db.execute("SELECT project_id FROM projects WHERE project_name=?", project_name)[0]['project_id']
        cutoff_date = request.form.get("cutoff")
        invoice = request.form.get("invoice")
        accrual_balance = request.form.get("accrual_balance")
        cash = request.form.get("cash")
        cash_balance = request.form.get("cash_balance")
        input_date = datetime.now()

        
        list = db.execute("SELECT project_id, cutoff, invoice FROM data WHERE project_id=? AND cutoff=?", project_id, cutoff_date)

        # initiate update if there is no data on cut-off date
        if len(list) == 0:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        if list[0]['cutoff'] != cutoff_date:
            errmsg = 1
            flash ("No update initiated for selected cut-off date, please initiate update")
            return render_template("uinit.html",errmsg=errmsg, projects=projects)

        # Check if already updated for selected cut-off date
        if list[0]['project_id'] == project_id and list[0]['cutoff'] == cutoff_date and list[0]['invoice'] != None:
            errmsg = 1
            flash ("Data already exists on the cut-off date selected")
            return render_template("ufinance.html",errmsg=errmsg, projects=projects)

        # Update entry
        db.execute("UPDATE data SET user_id = ?, invoice = ?, accrual_balance = ?, cash = ?,\
            cash_balance = ?, date = ? WHERE project_id = ? AND project_name = ? AND cutoff = ?",\
            user_id, invoice, accrual_balance, cash, cash_balance, input_date, project_id, project_name, cutoff_date)

        errmsg = 0
        flash("Update is successful")
        return render_template("ufinance.html",errmsg=errmsg, projects=projects)

    
    return render_template("ufinance.html", projects=projects)


@app.route("/uimport", methods=["GET", "POST"])
@login_required
def uimport():
    user_id = session.get("user_id")
    projects = db.execute("SELECT * FROM projects WHERE user_id=?", user_id)

    if request.method == "POST":
        # Ensure fields are not empty
        if not request.files["file"]:
            errmsg = 1
            flash("Please choose a file to import")
            return render_template("uimport.html", errmsg=errmsg)


        # Get file
        file = request.files["file"]
        #file.save(file.filename)

        # Read file and write in database
        wb = openpyxl.load_workbook(file)
        sheet = wb['import']
        for i in range(4, sheet.max_row+1):
            project_name = sheet.cell(row=i, column=1).value
            cutoff_date = sheet.cell(row=i, column=2).value
            start = sheet.cell(row=i, column=3).value
            contract_finish = sheet.cell(row=i, column=4).value
            planned_finish = sheet.cell(row=i, column=5).value
            total_mh = sheet.cell(row=i, column=6).value
            planned_mh = sheet.cell(row=i, column=7).value
            earned_mh = sheet.cell(row=i, column=8).value
            actual_mh = sheet.cell(row=i, column=9).value
            estimated_mh = sheet.cell(row=i, column=10).value
            bluecollar = sheet.cell(row=i, column=11).value
            whitecollar = sheet.cell(row=i, column=12).value
            subcon = sheet.cell(row=i, column=13).value
            income = sheet.cell(row=i, column=14).value
            expense = sheet.cell(row=i, column=15).value
            invoice = sheet.cell(row=i, column=16).value
            accrual_balance = sheet.cell(row=i, column=17).value
            cash = sheet.cell(row=i, column=18).value
            cash_balance = sheet.cell(row=i, column=19).value
            input_date = datetime.now()

            # Check if project exists in PROJECTS table
            name = db.execute("SELECT project_name FROM projects WHERE project_name=? AND user_id=?", project_name, user_id)
            if len(name) == 0:
                # Insert new row to PROJECTS table for new project
                db.execute("INSERT INTO projects(user_id, project_name, project_start, date) VALUES(?,?,?,?)",\
                user_id, project_name, date(start), input_date)

            project_id = db.execute("SELECT project_id FROM projects WHERE project_name=?", project_name)[0]['project_id']
            list = db.execute("SELECT user_id, project_id, cutoff FROM data WHERE user_id=? AND project_id=? AND cutoff=?", user_id, project_id, date(cutoff_date))

            # Create new entry in DATA table if there is no entry on cut-of date
            if len(list) == 0:
                db.execute("INSERT INTO data (user_id, project_id, project_name, cutoff, contract_finish, planned_finish,\
                total_mh, planned_mh, earned_mh, actual_mh, estimated_mh, bluecollar_mp, whitecollar_mp, subcon_mp, income, expense,\
                invoice, accrual_balance, cash, cash_balance, date) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",\
                user_id, project_id, project_name, date(cutoff_date), date(contract_finish), date(planned_finish),\
                total_mh, planned_mh, earned_mh, actual_mh, estimated_mh, bluecollar, whitecollar, subcon, income, expense, invoice, accrual_balance, cash, cash_balance, input_date)
                

            # Update entry on DATA tabel if there is entry on cut-of date
            if len(list) != 0:
                db.execute("UPDATE data SET  contract_finish=?, planned_finish=?,\
                total_mh=?, planned_mh=?, earned_mh=?, actual_mh=?, estimated_mh=?, bluecollar_mp=?, whitecollar_mp=?, subcon_mp=?, income=?, expense=?,\
                invoice=?, accrual_balance=?, cash=?, cash_balance=?, date=?\
                WHERE project_id = ? AND project_name = ? AND cutoff = ?",\
                date(contract_finish), date(planned_finish),\
                total_mh, planned_mh, earned_mh, actual_mh, estimated_mh, bluecollar, whitecollar, subcon, income, expense, invoice, accrual_balance, cash, cash_balance, input_date,\
                project_id, project_name, date(cutoff_date))

        errmsg = 0
        flash("Import is successful")
        return render_template("uimport.html",errmsg=errmsg)
   
    return render_template("uimport.html") 

# Admin Menu -----------------------------------

@app.route("/admprojects", methods=["GET", "POST"])
@login_required
def admprojects():

    projects = db.execute("SELECT * FROM projects")

    # Data Delete
    if request.method == "POST":
        project_id = request.form.get("project_id")
        db.execute("DELETE FROM projects WHERE project_id=?", project_id)
        errmsg = 0
        flash(f"project_name {project_id} is removed from database")
        return render_template("admprojects.html", projects=projects, errmsg=errmsg)

    projects = db.execute("SELECT * FROM projects")
    return render_template("admprojects.html", projects=projects)

@app.route("/admusers", methods=["GET", "POST"])
@login_required
def admusers():

    users = db.execute("SELECT * FROM users")

    # Data Delete
    if request.method == "POST":
        user = request.form.get("username")

        if user == 'admin':
            errmsg = 1
            flash("Admin can't be deleted")
            return render_template("admusers.html",users=users, errmsg=errmsg)
        else:
            db.execute("DELETE FROM users WHERE username=?", user)
            errmsg = 0
            flash(f"username {user} is removed from database")
            data = db.execute("SELECT * FROM users")
            return render_template("admusers.html", data=data, users=users, errmsg=errmsg)

    data = db.execute("SELECT * FROM users")
    return render_template("admusers.html",data=data, users=users)

@app.route("/admdata", methods=["GET", "POST"])
@login_required
def admdata():

    data = db.execute("SELECT * FROM data")

    # Data Delete
    if request.method == "POST":
        data_id = request.form.get("data_id")
        db.execute("DELETE FROM data WHERE id=?", data_id)
        errmsg = 0
        flash(f"data_id {data_id} is removed from database")
        return render_template("admdata.html",data=data, errmsg=errmsg)

    data = db.execute("SELECT * FROM data")
    return render_template("admdata.html",data=data)


# Others -----------------------------------
def format_server_time():
    server_time = time.localtime()
    return time.strftime("%I:%M:%S %p", server_time)
 
def number(value):
    """Format value as USD."""
    return f'{value:,.1f}'.replace(',', ' ')

def percent(value):
    """Format value as %"""
    return f"{value:.1f}%"

def date(value):
    """Format value as USD."""
    return str(value).removesuffix(" 00:00:00")

@app.route("/test", methods=["GET", "POST"])
@login_required
def test():
    return render_template("test.html")

